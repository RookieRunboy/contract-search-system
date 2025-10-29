"""客户分类映射加载器。

负责把 Excel 白名单映射加载到内存，并在文件变化时自动刷新，
供元数据提取阶段根据客户名称快速查找分类信息。
"""
from __future__ import annotations

import threading
from pathlib import Path
from typing import Dict, Optional, Tuple
import re

CustomerCategory = Tuple[Optional[str], Optional[str]]

# 预编译空白正则，便于在标准化名称时复用
_WHITESPACE_PATTERN = re.compile(r"\s+")
# 匹配甲乙丙丁方前缀，便于统一客户名称
_PARTY_PREFIX_PATTERN = re.compile(r"^[甲乙丙丁]方[:：\s]*")


def normalize_customer_key(value: Optional[str]) -> str:
    """将客户名称标准化为查找键。"""
    if not value:
        return ""

    text = str(value).strip()
    # 去除常见的全角空格，随后统一去掉所有空白
    text = text.replace("\u3000", " ")
    text = _PARTY_PREFIX_PATTERN.sub("", text)
    text = text.strip("（）()\u3000 ")
    # 对内部所有空白压缩并移除
    text = _WHITESPACE_PATTERN.sub("", text)
    if not text:
        return ""
    return text.casefold()


class CustomerCategoryLookup:
    """Excel 客户分类白名单的缓存查找器。"""

    def __init__(self, file_path: Optional[str]) -> None:
        self._file_path = Path(file_path).expanduser() if file_path else None
        self._mapping: Dict[str, CustomerCategory] = {}
        self._file_mtime: Optional[float] = None
        self._lock = threading.RLock()

    def lookup(self, customer_name: Optional[str]) -> CustomerCategory:
        """根据客户名称返回 (一级分类, 二级分类)。"""
        if not customer_name or not self._file_path:
            return (None, None)

        key = normalize_customer_key(customer_name)
        if not key:
            return (None, None)

        self._ensure_loaded()
        return self._mapping.get(key, (None, None))

    def refresh(self) -> None:
        """外部触发的强制刷新。"""
        if not self._file_path:
            return
        with self._lock:
            self._load_mapping(force=True)

    def _ensure_loaded(self) -> None:
        if not self._file_path:
            return
        with self._lock:
            self._load_mapping(force=False)

    def _load_mapping(self, force: bool) -> None:
        path = self._file_path
        if not path.exists():
            if self._mapping:
                print(f"客户分类白名单文件不存在: {path}")
                self._mapping.clear()
                self._file_mtime = None
            return

        mtime = path.stat().st_mtime
        if not force and self._file_mtime is not None and mtime <= self._file_mtime:
            return

        try:
            mapping = self._read_excel(path)
        except Exception as exc:  # pylint: disable=broad-exception-caught
            print(f"加载客户分类白名单失败: {exc}")
            return

        self._mapping = mapping
        self._file_mtime = mtime
        print(f"客户分类白名单已加载，共 {len(mapping)} 条记录")

    def _read_excel(self, path: Path) -> Dict[str, CustomerCategory]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pylint: disable=broad-exception-caught
            raise RuntimeError("缺少 openpyxl 依赖，无法读取客户分类白名单") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(min_row=1, values_only=True)
            try:
                header = next(rows)
            except StopIteration as exc:
                raise RuntimeError("客户分类白名单为空") from exc

            column_index = self._resolve_column_indices(header)
            if column_index["customer_name"] is None:
                raise RuntimeError("客户分类白名单缺少客户名称列")

            mapping: Dict[str, CustomerCategory] = {}
            duplicates: Dict[str, CustomerCategory] = {}
            for row in rows:
                name = self._get_cell_value(row, column_index["customer_name"])
                if not name:
                    continue

                level1 = self._get_cell_value(row, column_index["level1"])
                level2 = self._get_cell_value(row, column_index["level2"])

                key = normalize_customer_key(name)
                if not key:
                    continue

                record = (level1, level2)
                if key in mapping and mapping[key] != record:
                    duplicates[key] = mapping[key]
                    continue

                mapping[key] = record

            if duplicates:
                print(
                    f"客户分类白名单存在重复客户，已保留首条记录: {', '.join(duplicates.keys())}"
                )

            return mapping
        finally:
            workbook.close()

    @staticmethod
    def _resolve_column_indices(header_row: Tuple[Optional[object], ...]) -> Dict[str, Optional[int]]:
        normalized_headers = [CustomerCategoryLookup._normalize_header(cell) for cell in header_row]

        name_candidates = {"客户名称", "客户名", "客户", "公司名称", "名称"}
        level1_candidates = {"一级分类", "一级客户分类", "一级类", "一级"}
        level2_candidates = {"二级分类", "二级客户分类", "二级类", "二级"}

        def find_index(candidates: set[str]) -> Optional[int]:
            for idx, header in enumerate(normalized_headers):
                if header and header in candidates:
                    return idx
            return None

        return {
            "customer_name": find_index(name_candidates),
            "level1": find_index(level1_candidates),
            "level2": find_index(level2_candidates),
        }

    @staticmethod
    def _normalize_header(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.replace("\u3000", " ")
        text = _WHITESPACE_PATTERN.sub("", text)
        return text

    @staticmethod
    def _get_cell_value(row: Tuple[Optional[object], ...], index: Optional[int]) -> Optional[str]:
        if index is None or index >= len(row):
            return None
        value = row[index]
        if value is None:
            return None
        text = str(value).strip()
        return text or None
